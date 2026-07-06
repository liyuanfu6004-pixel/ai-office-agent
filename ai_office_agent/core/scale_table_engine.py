"""规模表智能识别引擎 — v1.2.1 优化（Sheet评分重构 + 多行表头 + 动态字段手动添加）。

v1.2.1 改进：
1. Sheet 评分重构：综合考虑 Sheet 名称权重（最高优先）+ 表格结构 + 字段关键词辅助
2. 多行表头识别：扫描前10行，支持合并单元格展开，多行组合表头自动合并
3. 动态字段改为手动添加：默认仅保留区县+点位名称，其余由用户手动选择
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .matcher import (
    any_match,
    is_match,
    match_sheet,
)
from ..utils.logger import setup_logger

logger = setup_logger()

# ====================================================================
# 1. Sheet 识别 —— 重构评分策略（v1.2.1）
# ====================================================================

# Sheet 名称关键词（命中即高分，最高权重）
_SHEET_NAME_SCALE_KEYWORDS: tuple[str, ...] = (
    "项目规模表", "工程规模表", "设计规模表",
    "规模表", "规模", "项目规模", "工程规模", "设计规模",
    "点位表", "站点表", "机房表",
)

# Sheet 名称反例关键词（命中即扣分）
_SHEET_NAME_ANTI_KEYWORDS: tuple[str, ...] = (
    "材料表", "物料表", "材料", "物料", "设备清单",
    "封面", "目录", "汇总", "说明", "备注", "模板",
    "勘察", "照片", "流程", "批复", "统计",
    "预算", "概算", "决算",
)

# 材料表典型特征：字段很少（≤6）、列名简单
_MATERIAL_KEYWORDS: tuple[str, ...] = (
    "材料名称", "物料名称", "规格", "型号", "数量", "单位",
    "单价", "合价", "材料编码", "物料编码",
    "名称", "规格型号",
)

# 规模表字段关键词（辅助评分，低权重）
_SCALE_SHEET_KEYWORDS: tuple[str, ...] = (
    "点位", "站点", "机房", "社区",
    "区县", "建设内容", "资源名称",
    "明细", "清单", "设备名称", "节点名称", "接入点",
    "光缆", "管道", "芯数", "长度", "地址",
    "经度", "纬度", "起点", "终点",
    "计划建设内容", "建设方案",
)


def score_sheet_likelihood(
    headers: list[str],
    data_rows: list[dict],
    sheet_name: str = "",
) -> float:
    """评估一个 Sheet 是规模表的可能性（0.0 ~ 1.0）—— v1.2.1 重构。

    评分维度（三层）：
    1. Sheet 名称命中规模表关键词（权重 0.45）——最高优先
    2. 表格结构特征（权重 0.35）——字段数、行数、列宽分布
    3. 字段关键词命中（权重 0.20）——辅助参考

    Returns:
        综合评分，越高越可能是规模表。
    """
    if not headers or not data_rows:
        return 0.0

    # ── 维度1：Sheet 名称评分 ──
    name_score = _score_sheet_name(sheet_name)

    # ── 维度2：表格结构评分 ──
    structure_score = _score_table_structure(headers, data_rows)

    # ── 维度3：字段关键词评分 ──
    keyword_score = _score_field_keywords(headers, data_rows)

    # 材料表惩罚：如果 Sheet 名称命中材料表关键词，大幅降权
    anti = _score_sheet_anti(sheet_name, headers)
    if anti >= 0.5:
        name_score *= 0.15

    total = round(
        name_score * 0.45 + structure_score * 0.35 + keyword_score * 0.20, 4
    )
    return min(total, 1.0)


def _score_sheet_name(sheet_name: str) -> float:
    """评分 Sheet 名称（0.0~1.0）。"""
    if not sheet_name:
        return 0.3  # 无名称时中性评分

    # 精确包含匹配
    for kw in _SHEET_NAME_SCALE_KEYWORDS:
        result = match_sheet(sheet_name, kw)
        if result.kind.name == "EXACT":
            return 1.0
        if result.is_match:
            # 根据关键词长度给分
            ratio = min(len(kw) / len(sheet_name), 1.0)
            return 0.8 + 0.2 * ratio

    # 模糊匹配检查
    best = 0.0
    for kw in _SHEET_NAME_SCALE_KEYWORDS:
        r = match_sheet(sheet_name, kw)
        if r.score > 60:
            best = max(best, r.score / 100.0)
    if best > 0:
        return best * 0.7  # 模糊匹配打折

    return 0.0


def _score_sheet_anti(sheet_name: str, headers: list[str]) -> float:
    """反例评分：如果是材料表/汇总表则返回高分。"""
    score = 0.0

    # Sheet 名称命中反例
    for kw in _SHEET_NAME_ANTI_KEYWORDS:
        r = match_sheet(sheet_name, kw)
        if r.is_match:
            score = max(score, r.score / 100.0)

    # 表头命中材料表关键词
    material_hits = 0
    for h in headers:
        for kw in _MATERIAL_KEYWORDS:
            if is_match(h, kw):
                material_hits += 1
                break
    if material_hits >= 3 and len(headers) <= 6:
        score = max(score, 0.7)

    return score


def _score_table_structure(headers: list[str], data_rows: list[dict]) -> float:
    """评估表格结构是否符合规模表特征（v1.2.1 新增）。

    规模表特征：
    - 字段较多（≥5 列）
    - 数据行较多（≥3 行）
    - 不是单列或两列表
    - 数据密度较高（非空单元格比例）
    """
    n_cols = len(headers)
    n_rows = len(data_rows)

    # 列数评分
    if n_cols >= 15:
        col_score = 1.0
    elif n_cols >= 10:
        col_score = 0.85
    elif n_cols >= 5:
        col_score = 0.6
    elif n_cols >= 3:
        col_score = 0.3
    else:
        col_score = 0.1  # 1-2 列不太可能是规模表

    # 行数评分
    if n_rows >= 20:
        row_score = 1.0
    elif n_rows >= 10:
        row_score = 0.8
    elif n_rows >= 3:
        row_score = 0.5
    else:
        row_score = 0.2

    # 数据密度评分（前 20 行非空单元格比例）
    sample = data_rows[:20]
    filled = 0
    total = len(headers) * len(sample)
    if total > 0:
        for row in sample:
            for h in headers:
                v = row.get(h)
                if v is not None and str(v).strip() != "":
                    filled += 1
    density = filled / total if total > 0 else 0.0

    return round(col_score * 0.4 + row_score * 0.3 + density * 0.3, 4)


def _score_field_keywords(headers: list[str], data_rows: list[dict]) -> float:
    """字段关键词命中评分（辅助，低权重）。

    v1.2.1：降低此维度权重为主评分服务。
    """
    if not headers:
        return 0.0

    hit_count = 0
    for h in headers:
        for kw in _SCALE_SHEET_KEYWORDS:
            if is_match(h, kw):
                hit_count += 1
                break

    # 命中率
    hit_rate = min(hit_count / max(len(headers), 1), 1.0)
    # 绝对命中数也加分
    abs_bonus = min(hit_count / 10.0, 1.0) * 0.5

    return round(hit_rate * 0.5 + abs_bonus, 4)


def detect_best_sheet(
    all_sheets: dict[str, tuple[list[str], list[dict]]],
) -> list[dict]:
    """分析全部 Sheet，按可能性评分排序。

    Args:
        all_sheets: {sheet_name: (headers, data_rows)}。

    Returns:
        排序后的候选列表，每项含 sheet_name / score / headers / data_rows / reason。
    """
    candidates: list[dict] = []
    for name, (headers, rows) in all_sheets.items():
        score = score_sheet_likelihood(headers, rows, sheet_name=name)
        candidates.append({
            "sheet_name": name,
            "score": score,
            "headers": headers,
            "data_rows": rows,
            "reason": _describe_score(name, score, headers, rows),
        })
    candidates.sort(key=lambda c: c["score"], reverse=True)
    logger.info(
        "Sheet 识别完成（v1.2.1 评分重构）：%d 张表，最佳 「%s」(%.3f)",
        len(candidates),
        candidates[0]["sheet_name"] if candidates else "无",
        candidates[0]["score"] if candidates else 0,
    )
    return candidates


def _describe_score(name: str, score: float, headers: list[str], rows: list[dict]) -> str:
    """生成评分描述文字。"""
    parts = []
    if score >= 0.7:
        parts.append("高概率规模表")
    elif score >= 0.4:
        parts.append("可能是规模表")
    else:
        parts.append("可能不是规模表")
    parts.append(f"({len(headers)} 列 × {len(rows)} 行")
    if name:
        parts.append(f"，Sheet名：{name}")
    parts.append(")")
    return "".join(parts)


# ====================================================================
# 2. 字段智能识别
# ====================================================================

# 点位名称关键词（按优先级排序）
_POINT_NAME_KEYWORDS: tuple[str, ...] = (
    "机房资管标准名称", "机房名称", "机房名",
    "站点名称", "站点名", "站点",
    "设备名称", "设备名",
    "资源名称", "资源名",
    "节点名称", "节点名", "节点",
    "社区名称", "社区名", "社区",
    "点位名称", "点位名", "点位",
    "管孔名称", "管道名称",
    "名称", "名字", "站名", "接入点名称",
)

# 区县关键词
_COUNTY_KEYWORDS: tuple[str, ...] = (
    "区县", "县区", "县", "区", "区域",
    "所属区县", "所在区县", "地区", "地市",
    "行政区", "行政区划",
)

# 起点关键词
_START_POINT_KEYWORDS: tuple[str, ...] = (
    "起点", "起点名称", "起始点", "起点站", "起始名称",
    "A端", "A端名称", "A端站点",
    "上游", "上游节点", "上游名称",
    "源端", "源端名称", "源节点",
    "起始", "起始站点", "开始节点",
    "上联", "上联站点", "上联点",
)

# 终点关键词
_END_POINT_KEYWORDS: tuple[str, ...] = (
    "终点", "终点名称", "终止点", "终点站", "终止名称",
    "Z端", "Z端名称", "Z端站点",
    "下游", "下游节点", "下游名称",
    "宿端", "宿端名称", "宿节点",
    "终止", "终止站点", "结束节点",
    "下联", "下联站点",
)


def _best_match(
    headers: list[str],
    keywords: tuple[str, ...],
) -> str | None:
    """在表头列表中找到第一个命中任一关键词的列名。"""
    for kw in keywords:
        result = any_match(kw, headers)
        if result is not None:
            return result.target_raw
    return None


def detect_point_name_field(headers: list[str]) -> str | None:
    """自动识别点位名称列。"""
    return _best_match(headers, _POINT_NAME_KEYWORDS)


def detect_county_field(headers: list[str]) -> str | None:
    """自动识别区县列。"""
    return _best_match(headers, _COUNTY_KEYWORDS)


def detect_start_field(headers: list[str]) -> str | None:
    """自动识别起点列。"""
    return _best_match(headers, _START_POINT_KEYWORDS)


def detect_end_field(headers: list[str]) -> str | None:
    """自动识别终点列。"""
    return _best_match(headers, _END_POINT_KEYWORDS)


def detect_all_fields(headers: list[str]) -> dict[str, str | None]:
    """一键识别全部关键字段。"""
    return {
        "point_name": detect_point_name_field(headers),
        "county": detect_county_field(headers),
        "start_point": detect_start_field(headers),
        "end_point": detect_end_field(headers),
    }


# ====================================================================
# 3. 点位生成规则
# ====================================================================

_CONCATENATE_PROJECT_TYPES: frozenset[str] = frozenset({
    "接入段", "城域网",
})


def should_concatenate(project_type: str | None) -> bool:
    """判断该项目类型是否应使用起点+终点拼接生成点位名称。"""
    return project_type in _CONCATENATE_PROJECT_TYPES


def generate_point_name(
    row: dict,
    point_field: str | None = None,
    start_field: str | None = None,
    end_field: str | None = None,
    use_concatenation: bool = False,
) -> str:
    """从数据行中提取/生成点位名称。"""
    if use_concatenation and start_field and end_field:
        start_val = _cell_text(row, start_field)
        end_val = _cell_text(row, end_field)
        if start_val and end_val:
            return f"{start_val}-{end_val}"
        if start_val:
            return start_val
        if end_val:
            return end_val
        return ""

    if point_field:
        return _cell_text(row, point_field)
    return ""


# ====================================================================
# 4. 动态字段 — v1.2.1 改为默认不导入，用户手动添加
# ====================================================================

# 已知概念关键词（用于自动标注已选字段的类型）
_KNOWN_DYNAMIC_KEYWORDS: dict[str, tuple[str, ...]] = {
    "长度": ("长度", "距离", "光缆长度", "电缆长度", "管道长度"),
    "芯数": ("芯数", "光缆芯数", "光纤芯数"),
    "经度": ("经度", "longitude", "LON"),
    "纬度": ("纬度", "latitude", "LAT"),
    "设备型号": ("设备型号", "型号", "设备规格", "规格型号"),
    "端口数": ("端口数", "端口数量", "接口数"),
    "带宽": ("带宽", "速率", "容量"),
    "建设方式": ("建设方式", "施工方式", "敷设方式"),
    "备注": ("备注", "说明", "描述", "备注说明"),
    "起点": ("起点", "起点名称", "起始点"),
    "终点": ("终点", "终点名称", "终止点"),
}


def classify_dynamic_fields(
    headers: list[str],
    occupied_fields: set[str],
) -> list[dict]:
    """列出所有可选的动态字段候选（v1.2.1：默认不导入，仅返回候选列表）。

    用户可在向导中手动选择要导入的字段。
    本函数返回全部非固定字段作为可选候选。

    Args:
        headers: 全部表头。
        occupied_fields: 已被分配为固定字段的列名（点位/区县/起点/终点）。

    Returns:
        每项含 name（原始列名）, label（已知概念标签或原始列名）, type, selected（默认 False）。
    """
    result: list[dict] = []
    for h in headers:
        if h in occupied_fields:
            continue

        label = h
        concept = _guess_dynamic_concept(h)
        if concept:
            label = concept

        result.append({
            "name": h,
            "label": label,
            "type": concept if concept else "other",
            "selected": False,  # v1.2.1：默认不选中
        })

    logger.debug("动态字段候选：%d 列（已排除 %d 个固定字段）",
                 len(result), len(occupied_fields))
    return result


def get_default_dynamic_fields() -> list[dict]:
    """返回默认的动态字段列表（v1.2.1：空，用户手动添加）。"""
    return []


def _guess_dynamic_concept(header: str) -> str | None:
    """尝试把表头归入已知概念。"""
    for concept, keywords in _KNOWN_DYNAMIC_KEYWORDS.items():
        if any_match(header, list(keywords)) is not None:
            return concept
    return None


# ====================================================================
# 5. 字段映射候选
# ====================================================================


@dataclass
class FieldMappingCandidates:
    """字段映射候选集合。"""
    point_name_candidates: list[str] = field(default_factory=list)
    county_candidates: list[str] = field(default_factory=list)
    start_point_candidates: list[str] = field(default_factory=list)
    end_point_candidates: list[str] = field(default_factory=list)

    point_name: str | None = None
    county: str | None = None
    start_point: str | None = None
    end_point: str | None = None


def build_field_candidates(headers: list[str]) -> FieldMappingCandidates:
    """从表头列表构建全部字段映射候选。"""
    candidates = FieldMappingCandidates()

    for h in headers:
        for kw in _POINT_NAME_KEYWORDS:
            if is_match(h, kw):
                candidates.point_name_candidates.append(h)
                break
        for kw in _COUNTY_KEYWORDS:
            if is_match(h, kw):
                candidates.county_candidates.append(h)
                break
        for kw in _START_POINT_KEYWORDS:
            if is_match(h, kw):
                candidates.start_point_candidates.append(h)
                break
        for kw in _END_POINT_KEYWORDS:
            if is_match(h, kw):
                candidates.end_point_candidates.append(h)
                break

    if candidates.point_name_candidates:
        candidates.point_name = candidates.point_name_candidates[0]
    if candidates.county_candidates:
        candidates.county = candidates.county_candidates[0]
    if candidates.start_point_candidates:
        candidates.start_point = candidates.start_point_candidates[0]
    if candidates.end_point_candidates:
        candidates.end_point = candidates.end_point_candidates[0]

    return candidates


# ====================================================================
# 6. 预览数据生成
# ====================================================================


def build_preview_rows(
    data_rows: list[dict],
    mapping: dict,
    dynamic_fields: list[dict],
    use_concatenation: bool,
    preview_count: int = 10,
) -> list[dict]:
    """按当前映射生成预览行。

    v1.2.1：仅预览用户已选中的动态字段。
    """
    rows = data_rows[:preview_count]
    result: list[dict] = []
    for row in rows:
        item: dict[str, Any] = {}
        county_field = mapping.get("county")
        item["county"] = _cell_text(row, county_field) if county_field else ""
        item["point_name"] = generate_point_name(
            row,
            point_field=mapping.get("point_name"),
            start_field=mapping.get("start_point"),
            end_field=mapping.get("end_point"),
            use_concatenation=use_concatenation,
        )
        for df in dynamic_fields:
            item[df["name"]] = _cell_text(row, df["name"])
        result.append(item)
    return result


# ====================================================================
# 7. Excel 多 Sheet 读取（v1.2.1 多行表头支持）
# ====================================================================


def read_all_sheets(path: str | Path) -> dict[str, tuple[list[str], list[dict]]]:
    """读取 Excel 的所有 Sheet（v1.2.1 升级：支持多行合并表头）。

    表头识别策略（v1.2.1 重写，正确处理三层合并表头）：
    1. 用 openpyxl merged_cells 构建「合并单元格」映射
    2. 扫描前 15 行，按"文本填充率 × 文本占比 × 非空列数"评分找表头起始行
    3. 向下扩展表头区域（连续有文本的行），确定表头结束行
    4. 按列构建复合表头：对每列从上到下，若该单元格在合并范围内则继承上层值，
       组合为"上层-中层-下层"，去重
    5. 数据行从表头结束行 + 1 开始
    6. 禁止生成 __col_x__ 占位名，兜底使用"列N"

    Args:
        path: .xlsx 文件路径。

    Returns:
        {sheet_name: (headers, data_rows)}。
    """
    import openpyxl

    wb = openpyxl.load_workbook(filename=str(path), read_only=False, data_only=True)
    result: dict[str, tuple[list[str], list[dict]]] = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            merged_ranges = list(ws.merged_cells.ranges)

            # 读取全部行的原始值（Cell.value）
            all_rows: list[list] = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            if not all_rows:
                result[sheet_name] = ([], [])
                continue

            max_cols = max((len(r) for r in all_rows), default=0)

            # 取前 15 行用于表头识别
            scan_limit = min(15, len(all_rows))
            scan_rows: list[list[str | None]] = []
            for r in all_rows[:scan_limit]:
                vals: list[str | None] = []
                for c in range(max_cols):
                    v = r[c] if c < len(r) else None
                    vals.append(str(v).strip() if v is not None else None)
                scan_rows.append(vals)

            # 构建合并单元格查询结构：(row_idx, col_idx) -> merged_range
            merged_lookup = _build_merged_lookup(merged_ranges)

            # 识别表头起始行和结束行
            header_start, header_end = _detect_header_region(
                scan_rows, max_cols, merged_lookup
            )

            # 按列构建复合表头（正确处理纵向合并继承）
            headers = _build_composite_headers(
                scan_rows, header_start, header_end, max_cols, merged_lookup
            )

            # 数据行从 header_end + 1 开始
            data_start = header_end + 1

            data_rows: list[dict] = []
            for r_idx in range(data_start, len(all_rows)):
                r = all_rows[r_idx]
                if all(v is None or (isinstance(v, str) and str(v).strip() == "") for v in r):
                    continue
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    row_dict[header] = r[col_idx] if col_idx < len(r) else None
                data_rows.append(row_dict)

            result[sheet_name] = (headers, data_rows)
            logger.info(
                "Sheet「%s」读取完成（v1.2.1）：表头行=%d~%d，%d 列，%d 行数据，合并区域=%d",
                sheet_name, header_start + 1, header_end + 1,
                len(headers), len(data_rows), len(merged_ranges),
            )

    finally:
        wb.close()

    logger.info("多 Sheet 读取完成（v1.2.1 多行表头）：%d 张表", len(result))
    return result


def _build_merged_lookup(merged_ranges) -> dict:
    """构建 (row_idx, col_idx) -> merged_range 的查询映射。

    row_idx / col_idx 为 0 基。返回的 merged_range 含 min_row/max_row/min_col/max_col
    （openpyxl 用 1 基，这里统一转 0 基比较）。
    """
    lookup: dict[tuple[int, int], dict] = {}
    for mr in merged_ranges:
        info = {
            "min_row": mr.min_row - 1,
            "max_row": mr.max_row - 1,
            "min_col": mr.min_col - 1,
            "max_col": mr.max_col - 1,
        }
        for r in range(info["min_row"], info["max_row"] + 1):
            for c in range(info["min_col"], info["max_col"] + 1):
                lookup[(r, c)] = info
    return lookup


def _get_merged_value(
    scan_rows: list[list[str | None]],
    row_idx: int,
    col_idx: int,
    merged_lookup: dict,
) -> str | None:
    """获取单元格的"真实"值（若是合并单元格，取左上角值）。

    这是关键：openpyxl 中合并单元格只有左上角有值，其他是 None。
    本函数让任意 (row, col) 都能取到所属合并范围的左上角值。
    """
    if (row_idx, col_idx) in merged_lookup:
        info = merged_lookup[(row_idx, col_idx)]
        # 取左上角
        if info["min_row"] < len(scan_rows) and info["min_col"] < len(scan_rows[info["min_row"]]):
            return scan_rows[info["min_row"]][info["min_col"]]
        return None
    # 非合并单元格，直接取值
    if row_idx < len(scan_rows) and col_idx < len(scan_rows[row_idx]):
        return scan_rows[row_idx][col_idx]
    return None


def _is_merged_cell(row_idx: int, col_idx: int, merged_lookup: dict) -> bool:
    """判断单元格是否在合并范围内。"""
    return (row_idx, col_idx) in merged_lookup


def _is_merged_anchor(row_idx: int, col_idx: int, merged_lookup: dict) -> bool:
    """判断单元格是否是合并范围的左上角（锚点）。"""
    if (row_idx, col_idx) not in merged_lookup:
        return False
    info = merged_lookup[(row_idx, col_idx)]
    return info["min_row"] == row_idx and info["min_col"] == col_idx


def _detect_header_region(
    scan_rows: list[list[str | None]],
    max_cols: int,
    merged_lookup: dict,
) -> tuple[int, int]:
    """识别表头区域（起始行，结束行），均 0 基。

    v1.2.1 策略：
    1. 跳过纯标题行（只有 1 个非空单元格，或是一个大合并单元格占满整行）
    2. 找到第一行有多个非空文本列的行作为表头起始
    3. 向下连续扩展表头区域，直到出现"数据行"特征（多数值是数字或行格式变化）
    """
    if not scan_rows:
        return 0, 0

    # 第一步：找到表头起始行
    header_start = 0
    for idx, row in enumerate(scan_rows):
        # 计算非空非合并覆盖的单元格数（独立列数）
        distinct_cols = 0
        text_cells = 0
        for c in range(min(len(row), max_cols)):
            v = row[c]
            if v and str(v).strip():
                # 是否是合并单元格的锚点
                if _is_merged_anchor(idx, c, merged_lookup):
                    distinct_cols += 1
                    text_cells += 1
                elif not _is_merged_cell(idx, c, merged_lookup):
                    # 非合并单元格，且非空
                    distinct_cols += 1
                    text_cells += 1

        # 标题行特征：只有 1 个独立文本列（且可能是个大合并）
        if text_cells <= 1:
            continue

        # 找到第一个有多列文本的行
        if distinct_cols >= 2:
            header_start = idx
            break

    # 第二步：向下扩展表头区域
    header_end = header_start
    for idx in range(header_start + 1, len(scan_rows)):
        row = scan_rows[idx]
        # 统计该行的文本单元格
        text_count = 0
        numeric_count = 0
        for c in range(min(len(row), max_cols)):
            v = row[c]
            if v is None:
                continue
            v_str = str(v).strip()
            if not v_str:
                continue
            # 判断是文本还是数字
            if _is_numeric(v_str):
                numeric_count += 1
            else:
                text_count += 1

        # 数据行特征：数字单元格多于文本单元格
        if numeric_count > text_count and numeric_count >= 2:
            break
        # 表头行特征：有文本单元格
        if text_count >= 1:
            header_end = idx
            continue
        # 空行：可能是表头结束
        if text_count == 0 and numeric_count == 0:
            # 检查下一行是否是数据，是则停止
            if idx + 1 < len(scan_rows):
                next_row = scan_rows[idx + 1]
                next_numeric = sum(
                    1 for v in next_row if v and _is_numeric(str(v).strip())
                )
                if next_numeric >= 2:
                    break
            continue

    return header_start, header_end


def _is_numeric(s: str) -> bool:
    """判断字符串是否是数字（含小数、负号、百分号）。"""
    s = s.strip().rstrip("%")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _build_composite_headers(
    scan_rows: list[list[str | None]],
    header_start: int,
    header_end: int,
    max_cols: int,
    merged_lookup: dict,
) -> list[str]:
    """按列构建复合表头（v1.2.1 核心算法）。

    对每一列，从 header_start 到 header_end 逐行向下：
    - 若该单元格是合并单元格的锚点 → 取其值作为一个表头层
    - 若该单元格被**纵向合并**覆盖（同列内 min_col == col 但 min_row != row）→ 锚点值
      已在上层加入，跳过
    - 若该单元格被**横向合并**覆盖（同行内 min_col != col）→ 继承该横向合并锚点
      的值作为一个表头层（去重，避免与同列已加入的相同值重复）
    - 若该单元格是非合并的有值单元格 → 取其值作为一个表头层
    - 去重后用 "-" 拼接

    例：
      H 列：第3行="社区前期建设情况"(H3:I3横向合并锚点) → 加入"社区前期建设情况"
            第4行="末端分光器端口数"(H4:H5纵向合并锚点) → 加入"末端分光器端口数"
            第5行被H4:H5纵向覆盖 → 跳过
      → 表头 = "社区前期建设情况-末端分光器端口数"

      I 列：第3行被H3:I3横向覆盖 → 继承"社区前期建设情况"
            第4行="发展用户数"(I4:I5纵向合并锚点) → 加入"发展用户数"
      → 表头 = "社区前期建设情况-发展用户数"

      A 列：第3行="序号"(A3:A5纵向合并锚点) → 加入"序号"
            第4、5行被纵向覆盖 → 跳过
      → 表头 = "序号"
    """
    headers: list[str] = []
    for col in range(max_cols):
        parts: list[str] = []
        seen_values: set[str] = set()

        for row_idx in range(header_start, header_end + 1):
            if row_idx >= len(scan_rows):
                continue
            if col >= len(scan_rows[row_idx]):
                continue

            # 判断合并状态
            if (row_idx, col) in merged_lookup:
                info = merged_lookup[(row_idx, col)]
                is_anchor = (
                    info["min_row"] == row_idx and info["min_col"] == col
                )
                is_vertical_cover = (
                    info["min_col"] == col and info["min_row"] != row_idx
                )
                is_horizontal_cover = (info["min_col"] != col)

                if is_anchor:
                    # 锚点：取自己的值
                    value = scan_rows[row_idx][col]
                elif is_vertical_cover:
                    # 纵向合并覆盖：锚点在上层同列，值已加入，跳过
                    continue
                elif is_horizontal_cover:
                    # 横向合并覆盖：继承同行锚点的值
                    anchor_row = info["min_row"]
                    anchor_col = info["min_col"]
                    if anchor_row < len(scan_rows) and anchor_col < len(scan_rows[anchor_row]):
                        value = scan_rows[anchor_row][anchor_col]
                    else:
                        continue
                else:
                    continue
            else:
                # 非合并单元格
                value = scan_rows[row_idx][col]

            if value is None:
                continue
            value = str(value).strip()
            if not value:
                continue

            if value not in seen_values:
                parts.append(value)
                seen_values.add(value)

        if not parts:
            headers.append(f"列{col + 1}")
        elif len(parts) == 1:
            headers.append(parts[0])
        else:
            headers.append("-".join(parts))

    return headers


# ====================================================================
# 8. 工具
# ====================================================================


def _cell_text(row: dict, header: str | None) -> str:
    """从行字典取值并转为干净文本。"""
    if not header:
        return ""
    v = row.get(header)
    if v is None:
        return ""
    return str(v).strip()


# ====================================================================
# 9. 规模表记录构建（v1.2.1：仅包含用户选中的动态字段）
# ====================================================================


def build_point_records(
    data_rows: list[dict],
    mapping: dict,
    dynamic_fields: list[dict],
    use_concatenation: bool,
) -> list[dict]:
    """将 Excel 数据行转为 point_dictionary 表记录。

    v1.2.1：仅导入用户手动选中的动态字段。
    """
    records: list[dict] = []
    for row in data_rows:
        point_name = generate_point_name(
            row,
            point_field=mapping.get("point_name"),
            start_field=mapping.get("start_point"),
            end_field=mapping.get("end_point"),
            use_concatenation=use_concatenation,
        )
        if not point_name:
            continue

        county_field = mapping.get("county")
        county = _cell_text(row, county_field) if county_field else ""

        if use_concatenation and mapping.get("start_point") and mapping.get("end_point"):
            s = _cell_text(row, mapping["start_point"])
            e = _cell_text(row, mapping["end_point"])
            original = f"{s}-{e}" if s and e else (s or e)
        else:
            original = _cell_text(row, mapping.get("point_name"))

        dynamic_data: dict[str, str] = {}
        for df in dynamic_fields:
            dynamic_data[df["name"]] = _cell_text(row, df["name"])

        records.append({
            "standard_point_name": point_name,
            "county": county,
            "original_name": original,
            "dynamic_data": dynamic_data,
        })

    logger.info("规模表记录构建：%d 条", len(records))
    return records
