"""扫描结果中心页面 — v1.2.2 升级（人工确认 + 重新匹配 + 学习机制 + 批量确认 + 导出）。

扫描结果中心是点位匹配与文件状态分析的核心视图。

v1.2.2 新增：
    - 人工确认：结果列表增加"确认"按钮列，状态切换"未确认"/"已确认"
    - 重新匹配：NOT_FOUND / MULTIPLE_MATCH / PARTIAL_MATCH 可重新选择目录
    - 学习机制：确认结果保存到 scan_match_history 表，下次扫描优先使用
    - 批量确认：全部确认 / 批量确认已匹配项
    - 导出 Excel：含标准点位/实际目录/CAD状态/预算状态/匹配率/建议/确认状态
    - 约束：所有操作仅修改扫描结果和确认历史，禁止修改文件/目录
"""
from __future__ import annotations

import time
from datetime import datetime, timezone

from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ....config import AppConfig
from ....core import (
    projects_repository,
    scan_match_history_repository,
)
from ....core.database import Database
from ....core.scan_result import (
    MatchStatus,
    ScanResultItem,
    ScanResultSummary,
)
from ....utils.logger import setup_logger
from .base_page import BasePage

logger = setup_logger()

# ====================================================================
# 颜色常量
# ====================================================================

_MATCH_COLORS: dict[MatchStatus, QColor] = {
    MatchStatus.MATCHED: QColor("#107C10"),
    MatchStatus.PARTIAL_MATCH: QColor("#FFB900"),
    MatchStatus.NOT_FOUND: QColor("#D13438"),
    MatchStatus.MULTIPLE_MATCH: QColor("#FF8C00"),
}

# 筛选占位项
_FILTER_ALL = "全部"


# ====================================================================
# 预留接口（仅定义签名，禁止实现）
# ====================================================================


class RenamePreviewInterface:
    """重命名预览接口（预留）。

    未来用于预览文件夹重命名为标准点位名称的效果。
    本版仅定义接口，不实现。
    """

    @staticmethod
    def preview_rename(scan_result: ScanResultItem) -> dict:
        """预览重命名效果。

        Returns:
            {"old_name": str, "new_name": str, "conflicts": list[str]}
        """
        raise NotImplementedError("v1.2 预留，不实现")


class FolderBuilderInterface:
    """文件夹构建接口（预留）。

    未来用于自动创建缺失的点位文件夹结构。
    本版仅定义接口，不实现。
    """

    @staticmethod
    def build_folder_structure(scan_results: list[ScanResultItem]) -> dict:
        """生成文件夹构建计划。

        Returns:
            {"to_create": list[dict], "to_move": list[dict], "dry_run": bool}
        """
        raise NotImplementedError("v1.2 预留，不实现")


class HealthScoreInterface:
    """健康评分接口（预留）。

    未来用于根据匹配率、文件完整度等计算项目健康分。
    本版仅定义接口，不实现。
    """

    @staticmethod
    def calculate_health(scan_results: list[ScanResultItem]) -> dict:
        """计算项目健康评分。

        Returns:
            {"score": float, "grade": str, "factors": list[dict]}
        """
        raise NotImplementedError("v1.2 预留，不实现")


class AISuggestionInterface:
    """AI 建议接口（预留）。

    未来用于根据扫描结果生成 AI 建议。
    本版仅定义接口，不实现。
    """

    @staticmethod
    def generate_suggestions(scan_result: ScanResultItem) -> list[str]:
        """生成 AI 建议。

        Returns:
            建议文本列表。
        """
        raise NotImplementedError("v1.2 预留，不实现")


# ====================================================================
# v1.4.1：文件整理预览对话框（可调整大小 + 滚动）
# ====================================================================


class OrganizePreviewDialog(QDialog):
    """文件整理预览对话框。

    特点：
    - 可调整大小，带右下角尺寸手柄
    - 内容区域支持垂直/水平滚动
    - 显示完整分类明细，不做数量截断
    - 底部「整理」+「取消」按钮，点击整理后执行实际移动
    """

    def __init__(self, plan, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("整理文件预览")
        self.setMinimumSize(720, 480)
        self.resize(960, 640)
        self.setSizeGripEnabled(True)
        self._setup_ui(plan)

    def _setup_ui(self, plan) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        self.text_edit = QPlainTextEdit(self)
        self.text_edit.setReadOnly(True)
        self.text_edit.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
        self.text_edit.setPlainText(self._build_text(plan))

        # 等宽字体，方便路径对齐
        font = self.text_edit.font()
        font.setFamily("Consolas")
        font.setPointSize(10)
        self.text_edit.setFont(font)

        layout.addWidget(self.text_edit)

        # 自定义按钮：取消 + 整理
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("取消", self)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        organize_btn = QPushButton("整理", self)
        organize_btn.setDefault(True)
        organize_btn.setStyleSheet(
            "QPushButton { background: #107C10; color: white; border: none; "
            "border-radius: 4px; padding: 6px 20px; font-size: 13px; font-weight: bold; }"
            "QPushButton:hover { background: #0E6A0E; }"
        )
        organize_btn.clicked.connect(self.accept)
        btn_layout.addWidget(organize_btn)

        layout.addLayout(btn_layout)

    def _build_text(self, plan) -> str:
        """根据整理计划构建完整预览文本。"""
        lines = [
            "=== 整理文件预览 ===",
            f"项目：{plan.project_path}",
            f"总文件：{plan.total_files}（图纸={plan.drawing_count} 预算={plan.budget_count} 其他={plan.other_count}）",
            f"点位：{len(plan.points)} 个",
        ]

        if plan.conflicts:
            lines.append(f"\n冲突文件：{len(plan.conflicts)}")
            for c in plan.conflicts:
                lines.append(f"  - {c}")

        lines.append("\n--- 分类明细 ---")
        for pname, items in plan.points.items():
            lines.append(f"\n【{pname}】")
            for r in items:
                lines.append(f"  {r.file.file_name} → {r.category}（{r.reason}）")

        return "\n".join(lines)



class StatCard(QFrame):
    """单个统计指标卡片（Windows 11 风格）。"""

    def __init__(
        self,
        title: str,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setObjectName("StatCard")
        self.setMinimumSize(100, 80)
        self.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed
        )
        self.setStyleSheet("""
            StatCard {
                background: #FFFFFF;
                border: 1px solid #EAEAEA;
                border-radius: 8px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(4)

        self.title_label = QLabel(title, self)
        self.title_label.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
        layout.addWidget(self.title_label)

        self.value_label = QLabel("0", self)
        self.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #1B1B1B; background: transparent;"
        )
        layout.addWidget(self.value_label)

    def set_value(self, value: int) -> None:
        """更新统计数值。"""
        self.value_label.setText(str(value))


# ====================================================================
# 统计卡片组
# ====================================================================


class StatCardRow(QWidget):
    """统计卡片行：总点位/已匹配/部分匹配/未匹配/CAD缺失/预算缺失/已确认（v1.2.2）。"""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        self.total_card = StatCard("总点位")
        self.matched_card = StatCard("已匹配")
        self.partial_card = StatCard("部分匹配")
        self.not_found_card = StatCard("未匹配")
        self.cad_missing_card = StatCard("CAD 缺失")
        self.budget_missing_card = StatCard("预算缺失")
        self.confirmed_card = StatCard("已确认")

        layout.addWidget(self.total_card)
        layout.addWidget(self.matched_card)
        layout.addWidget(self.partial_card)
        layout.addWidget(self.not_found_card)
        layout.addWidget(self.cad_missing_card)
        layout.addWidget(self.budget_missing_card)
        layout.addWidget(self.confirmed_card)

        # 颜色标注
        self.matched_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #107C10; background: transparent;"
        )
        self.partial_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #FFB900; background: transparent;"
        )
        self.not_found_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #D13438; background: transparent;"
        )
        self.cad_missing_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #D13438; background: transparent;"
        )
        self.budget_missing_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #E74856; background: transparent;"
        )
        self.confirmed_card.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #0067C0; background: transparent;"
        )

    def update_stats(self, summary: ScanResultSummary) -> None:
        """根据汇总更新全部卡片（v1.2.2 含已确认）。"""
        self.total_card.set_value(summary.total_points)
        self.matched_card.set_value(summary.matched_count)
        self.partial_card.set_value(summary.partial_match_count)
        self.not_found_card.set_value(summary.not_found_count)
        self.cad_missing_card.set_value(summary.cad_missing_count)
        self.budget_missing_card.set_value(summary.budget_missing_count)
        self.confirmed_card.set_value(summary.confirmed_count)


# ====================================================================
# 结果列表表格
# ====================================================================

# 固定列定义（v1.2.2 增加"确认"列）
_SCAN_COLUMNS: list[tuple[str, int]] = [
    ("状态", 80),
    ("标准点位", 200),
    ("实际文件夹", 170),
    ("匹配率", 72),
    ("CAD", 56),
    ("预算", 56),
    ("建议", 220),
    ("确认", 60),
]

# 列索引常量
_COL_STATUS = 0
_COL_STANDARD = 1
_COL_FOLDER = 2
_COL_SCORE = 3
_COL_CAD = 4
_COL_BUDGET = 5
_COL_SUGGESTION = 6
_COL_CONFIRM = 7


class ScanResultTable(QTableWidget):
    """扫描结果列表表格（v1.2.2 升级：增加确认按钮列）。

    固定 8 列：状态/标准点位/实际文件夹/匹配率/CAD/预算/建议/确认。
    支持排序、筛选、搜索。
    禁止修改数据。
    """

    # 双击结果行时发出，携带 ScanResultItem
    item_selected = Signal(object)
    # 确认状态切换时发出，携带 (row_index, ScanResultItem)
    confirm_toggled = Signal(int, object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._items: list[ScanResultItem] = []
        self._setup_table()
        self.cellDoubleClicked.connect(self._on_double_click)

    def _setup_table(self) -> None:
        """设置表格属性。"""
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setAlternatingRowColors(True)
        self.verticalHeader().setVisible(False)
        self.setSortingEnabled(True)

        header = self.horizontalHeader()
        header.setSectionsClickable(True)

        # 设置列（最后列固定宽度，不拉伸）
        titles = [c[0] for c in _SCAN_COLUMNS]
        self.setColumnCount(len(titles))
        self.setHorizontalHeaderLabels(titles)
        for col, w in enumerate(_SCAN_COLUMNS):
            self.setColumnWidth(col, w[1])
        # v1.4.1 修复：建议列改为可交互调整，允许拖动列宽；确认列固定
        header.setStretchLastSection(False)
        header.setSectionResizeMode(_COL_SUGGESTION, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(_COL_CONFIRM, QHeaderView.ResizeMode.Fixed)
        # 表格内容超出时显示滚动条，避免建议列被压缩
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

    def load_results(self, items: list[ScanResultItem]) -> None:
        """加载扫描结果列表。

        Args:
            items: ScanResultItem 列表。
        """
        self._items = items
        self.setSortingEnabled(False)
        self.setRowCount(len(items))
        for row, item in enumerate(items):
            self._fill_row(row, item)
        self.setSortingEnabled(True)

    def _fill_row(self, row: int, item: ScanResultItem) -> None:
        """填充一行数据（v1.2.2：增加确认按钮）。"""
        # 状态（带颜色）
        status_text = item.match_status.label
        status_item = QTableWidgetItem(status_text)
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        color = _MATCH_COLORS.get(item.match_status, QColor("#8A8A8A"))
        status_item.setForeground(color)
        font = status_item.font()
        font.setBold(True)
        status_item.setFont(font)
        status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.setItem(row, _COL_STATUS, status_item)

        # 标准点位
        self._set_text_cell(row, _COL_STANDARD, item.standard_point_name)

        # 实际文件夹
        folder_text = item.matched_folder or "—"
        self._set_text_cell(row, _COL_FOLDER, folder_text)

        # 匹配率
        score_text = f"{item.match_percent}%" if item.matched_folder else "—"
        score_item = QTableWidgetItem(score_text)
        score_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        score_item.setFlags(score_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.setItem(row, _COL_SCORE, score_item)

        # CAD 状态
        cad_item = QTableWidgetItem(item.cad_status)
        cad_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if item.cad_status == "有":
            cad_item.setForeground(QColor("#107C10"))
        else:
            cad_item.setForeground(QColor("#D13438"))
        cad_item.setFlags(cad_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.setItem(row, _COL_CAD, cad_item)

        # 预算状态
        budget_item = QTableWidgetItem(item.budget_status)
        budget_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        if item.budget_status == "有":
            budget_item.setForeground(QColor("#107C10"))
        else:
            budget_item.setForeground(QColor("#D13438"))
        budget_item.setFlags(budget_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.setItem(row, _COL_BUDGET, budget_item)

        # 建议
        suggestion_text = item.suggestion or "—"
        self._set_text_cell(row, _COL_SUGGESTION, suggestion_text)

        # v1.2.2：确认按钮
        confirm_btn = QPushButton(item.confirmed_label)
        confirm_btn.setFixedSize(58, 26)
        if item.confirmed:
            confirm_btn.setStyleSheet(
                "QPushButton { background: #107C10; color: white; border: none; "
                "border-radius: 4px; font-size: 12px; }"
                "QPushButton:hover { background: #0E6A0E; }"
            )
        else:
            confirm_btn.setStyleSheet(
                "QPushButton { background: #FAFAFA; color: #8A8A8A; "
                "border: 1px solid #D0D0D0; border-radius: 4px; font-size: 12px; }"
                "QPushButton:hover { background: #E5F1FB; color: #0067C0; border-color: #0067C0; }"
            )
        confirm_btn.clicked.connect(
            lambda checked, r=row, it=item: self.confirm_toggled.emit(r, it)
        )
        self.setCellWidget(row, _COL_CONFIRM, confirm_btn)

    def _set_text_cell(self, row: int, col: int, text: str) -> None:
        """设置文本单元格。"""
        item = QTableWidgetItem(text)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.setItem(row, col, item)

    def apply_filter(
        self,
        status_filter: str | None = None,
        search_text: str = "",
        cad_filter: str | None = None,
        budget_filter: str | None = None,
    ) -> None:
        """按条件筛选行。

        Args:
            status_filter: 匹配状态筛选（"已匹配"/"部分匹配"/"未找到"/"多候选" 或 None=全部）。
            search_text: 搜索文本（匹配标准点位名称或实际文件夹名）。
            cad_filter: CAD 状态筛选（"有"/"无" 或 None=全部）。
            budget_filter: 预算状态筛选（"有"/"无" 或 None=全部）。
        """
        search_lower = search_text.lower().strip()
        for row in range(self.rowCount()):
            item = self._items[row] if row < len(self._items) else None
            visible = True

            if status_filter and item:
                if item.match_status.label != status_filter:
                    visible = False

            if search_lower:
                name_match = (
                    item.standard_point_name.lower().find(search_lower) >= 0
                    if item else False
                )
                folder_match = (
                    (item.matched_folder or "").lower().find(search_lower) >= 0
                    if item else False
                )
                if not name_match and not folder_match:
                    visible = False

            if cad_filter and item:
                if item.cad_status != cad_filter:
                    visible = False

            if budget_filter and item:
                if item.budget_status != budget_filter:
                    visible = False

            self.setRowHidden(row, not visible)

    def _on_double_click(self, row: int, _col: int) -> None:
        """双击行 → 发出 item_selected 信号。"""
        if 0 <= row < len(self._items):
            self.item_selected.emit(self._items[row])


# ====================================================================
# 详情预览面板
# ====================================================================


class DetailPreviewPanel(QFrame):
    """右侧详情预览面板。

    双击结果列表中的行后，在此面板显示点位的详细信息。
    仅查看，禁止修改。
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("DetailPreviewPanel")
        self.setMinimumWidth(280)
        self.setStyleSheet("""
            #DetailPreviewPanel {
                background: #FFFFFF;
                border: 1px solid #EAEAEA;
                border-radius: 8px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        # 标题
        title = QLabel("点位详情", self)
        title.setObjectName("PanelTitle")
        layout.addWidget(title)

        # 提示
        self.hint_label = QLabel("双击结果行查看详情", self)
        self.hint_label.setObjectName("PanelHint")
        layout.addWidget(self.hint_label)

        # 详情内容区域
        self.content_widget = QWidget(self)
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(6)
        self.content_widget.hide()
        layout.addWidget(self.content_widget)

        # 详情标签
        self.field_labels: dict[str, QLabel] = {}

        # 字段定义：(标签, 键)
        fields = [
            ("标准点位名称", "standard_point_name"),
            ("Excel 原始名称", "original_name"),
            ("实际匹配目录", "matched_folder"),
            ("匹配率", "match_percent"),
            ("匹配状态", "match_status"),
            ("CAD 文件数量", "cad_file_count"),
            ("预算文件数量", "budget_file_count"),
            ("CAD 状态", "cad_status"),
            ("预算状态", "budget_status"),
            ("建议说明", "suggestion"),
        ]
        for label_text, key in fields:
            row_layout = QHBoxLayout()
            row_layout.setSpacing(8)
            lbl = QLabel(label_text + "：", self.content_widget)
            lbl.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
            lbl.setFixedWidth(110)
            val = QLabel("—", self.content_widget)
            val.setStyleSheet("color: #1B1B1B; font-size: 12px; background: transparent;")
            val.setWordWrap(True)
            val.setTextInteractionFlags(
                Qt.TextInteractionFlag.TextSelectableByMouse
            )
            row_layout.addWidget(lbl)
            row_layout.addWidget(val, 1)
            self.content_layout.addLayout(row_layout)
            self.field_labels[key] = val

        # 扫描文件列表
        files_label = QLabel("扫描到的文件列表：", self.content_widget)
        files_label.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
        self.content_layout.addWidget(files_label)

        self.files_list = QLabel("—", self.content_widget)
        self.files_list.setStyleSheet(
            "color: #1B1B1B; font-size: 11px; background: #FAFBFC; "
            "border: 1px solid #ECECEC; border-radius: 4px; padding: 6px;"
        )
        self.files_list.setWordWrap(True)
        self.files_list.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse
        )
        self.content_layout.addWidget(self.files_list)

        layout.addStretch()

    def show_detail(self, item: ScanResultItem) -> None:
        """显示点位详情。"""
        self.hint_label.hide()
        self.content_widget.show()

        self.field_labels["standard_point_name"].setText(item.standard_point_name or "—")
        self.field_labels["original_name"].setText(item.original_name or "—")
        self.field_labels["matched_folder"].setText(item.matched_folder or "（未匹配）")
        self.field_labels["match_percent"].setText(
            f"{item.match_percent}%" if item.matched_folder else "—"
        )
        self.field_labels["match_status"].setText(item.match_status.label)
        # 匹配状态着色
        color = _MATCH_COLORS.get(item.match_status, QColor("#8A8A8A"))
        self.field_labels["match_status"].setStyleSheet(
            f"color: {color.name()}; font-size: 12px; font-weight: bold; background: transparent;"
        )
        self.field_labels["cad_file_count"].setText(str(item.cad_file_count))
        self.field_labels["budget_file_count"].setText(str(item.budget_file_count))
        self.field_labels["cad_status"].setText(item.cad_status)
        self.field_labels["budget_status"].setText(item.budget_status)
        self.field_labels["suggestion"].setText(item.suggestion or "—")

        # 文件列表
        if item.scanned_files:
            files_text = "\n".join(item.scanned_files[:30])
            if len(item.scanned_files) > 30:
                files_text += f"\n... 还有 {len(item.scanned_files) - 30} 个文件"
            self.files_list.setText(files_text)
        else:
            self.files_list.setText("（无扫描文件）")

    def clear_detail(self) -> None:
        """清空详情面板。"""
        self.hint_label.show()
        self.content_widget.hide()
        for lbl in self.field_labels.values():
            lbl.setText("—")


# ====================================================================
# 筛选栏
# ====================================================================


class ScanFilterBar(QWidget):
    """扫描结果筛选栏：状态 + 搜索 + CAD + 预算。"""

    filter_changed = Signal()

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        # 状态筛选
        layout.addWidget(QLabel("状态：", self))
        self.status_combo = QComboBox(self)
        self.status_combo.setMinimumWidth(100)
        self.status_combo.addItem(_FILTER_ALL)
        self.status_combo.addItems(["已匹配", "部分匹配", "未找到", "多候选"])
        layout.addWidget(self.status_combo)

        # 搜索框
        layout.addWidget(QLabel("搜索：", self))
        self.search_edit = QLineEdit(self)
        self.search_edit.setPlaceholderText("搜索点位名称或文件夹 ...")
        self.search_edit.setClearButtonEnabled(True)
        layout.addWidget(self.search_edit, 1)

        # CAD 筛选
        layout.addWidget(QLabel("CAD：", self))
        self.cad_combo = QComboBox(self)
        self.cad_combo.setMinimumWidth(70)
        self.cad_combo.addItem(_FILTER_ALL)
        self.cad_combo.addItems(["有", "无"])
        layout.addWidget(self.cad_combo)

        # 预算筛选
        layout.addWidget(QLabel("预算：", self))
        self.budget_combo = QComboBox(self)
        self.budget_combo.setMinimumWidth(70)
        self.budget_combo.addItem(_FILTER_ALL)
        self.budget_combo.addItems(["有", "无"])
        layout.addWidget(self.budget_combo)

        # 信号
        self.status_combo.currentIndexChanged.connect(self.filter_changed.emit)
        self.search_edit.textChanged.connect(self.filter_changed.emit)
        self.cad_combo.currentIndexChanged.connect(self.filter_changed.emit)
        self.budget_combo.currentIndexChanged.connect(self.filter_changed.emit)

    def filters(self) -> dict:
        """返回当前筛选条件。"""
        def combo_val(c: QComboBox) -> str | None:
            t = c.currentText()
            return None if t == _FILTER_ALL else t
        return {
            "status": combo_val(self.status_combo),
            "search": self.search_edit.text(),
            "cad": combo_val(self.cad_combo),
            "budget": combo_val(self.budget_combo),
        }


# ====================================================================
# 重新匹配对话框（v1.2.2 新增）
# ====================================================================


class RematchDialog(QDialog):
    """重新匹配对话框：列出候选目录供用户选择。"""

    def __init__(
        self,
        standard_point_name: str,
        candidates: list[str],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(f"重新匹配 —— {standard_point_name}")
        self.setMinimumSize(500, 350)
        self._selected_folder: str | None = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        # 说明
        hint = QLabel(
            f"点位「{standard_point_name}」的候选文件夹：\n"
            f"请选择一个文件夹作为匹配目标。此选择将保存到历史记录，下次扫描时优先使用。"
        )
        hint.setStyleSheet("color: #555;")
        layout.addWidget(hint)

        # 候选列表
        self._table = QTableWidget()
        self._table.setColumnCount(1)
        self._table.setHorizontalHeaderLabels(["候选文件夹"])
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setRowCount(len(candidates))
        for i, c in enumerate(candidates):
            self._table.setItem(i, 0, QTableWidgetItem(c))
        layout.addWidget(self._table, 1)

        # 按钮
        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.accepted.connect(self._on_accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _on_accept(self) -> None:
        """确认选择。"""
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请选择一个候选文件夹。")
            return
        item = self._table.item(row, 0)
        if item:
            self._selected_folder = item.text()
        self.accept()

    def selected_folder(self) -> str | None:
        """返回用户选择的文件夹名称。"""
        return self._selected_folder


# ====================================================================
# 扫描结果中心页面（v1.2.2 全面升级）
# ====================================================================


class ScanCenterPage(BasePage):
    """扫描结果中心页面（v1.2.2 升级：人工确认 + 重新匹配 + 学习 + 批量 + 导出）。

    布局：
        顶部：项目名称、扫描时间、扫描耗时、扫描目录
        按钮行：执行扫描 / 全部确认 / 批量确认已匹配 / 重新匹配 / 导出Excel
        统计卡片行（含已确认）
        结果列表 + 右侧详情预览面板
    """

    def __init__(
        self,
        config: AppConfig | None = None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(
            title="扫描中心",
            subtitle="点位匹配与文件状态分析",
            parent=parent,
        )
        self._config = config or AppConfig()
        self._project_id: int | None = None
        self._project_name: str = ""
        self._summary: ScanResultSummary | None = None
        self._has_valid_session: bool = False

        self._setup_project_selector()
        self._setup_info_bar()
        self._setup_stat_cards()
        self._setup_result_area()

        self._clear_display()

    # ------------------------------------------------------------------ 顶部信息栏

    def _setup_project_selector(self) -> None:
        """项目选择器：标题下方的独立横幅，用于选择已导入项目。"""
        self.project_selector_bar = QWidget(self)
        self.project_selector_bar.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed,
        )
        layout = QHBoxLayout(self.project_selector_bar)
        layout.setContentsMargins(0, 8, 0, 8)
        layout.setSpacing(12)

        project_label = QLabel("项目：", self.project_selector_bar)
        project_label.setStyleSheet(
            "color: #1B1B1B; font-size: 14px; font-weight: 600; background: transparent;"
        )
        layout.addWidget(project_label)

        self.project_combo = QComboBox(self.project_selector_bar)
        self.project_combo.setMinimumWidth(300)
        self.project_combo.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed,
        )
        self.project_combo.currentIndexChanged.connect(self._on_project_selected)
        layout.addWidget(self.project_combo, 1)

        layout.addStretch()
        self.content_layout.addWidget(self.project_selector_bar)

    def _setup_info_bar(self) -> None:
        """顶部信息栏：扫描时间/耗时/目录 + 操作按钮。"""
        bar = QWidget(self)
        bar.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        # 信息区域
        info_layout = QVBoxLayout()
        info_layout.setSpacing(2)

        meta_row = QHBoxLayout()
        meta_row.setSpacing(20)

        self.scan_time_label = QLabel("扫描时间：—", bar)
        self.scan_time_label.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
        meta_row.addWidget(self.scan_time_label)

        self.scan_duration_label = QLabel("耗时：—", bar)
        self.scan_duration_label.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
        meta_row.addWidget(self.scan_duration_label)

        self.scan_dir_label = QLabel("目录：—", bar)
        self.scan_dir_label.setStyleSheet("color: #8A8A8A; font-size: 12px; background: transparent;")
        meta_row.addWidget(self.scan_dir_label)

        meta_row.addStretch()
        info_layout.addLayout(meta_row)

        layout.addLayout(info_layout, 1)

        # 操作按钮（v1.2.2 新增确认/匹配/导出按钮）
        # v1.4.1：删除重复的「重新扫描」按钮，仅保留「执行扫描」
        self.scan_btn = QPushButton("执行扫描", bar)
        self.scan_btn.setDefault(True)
        self.scan_btn.clicked.connect(self._on_scan)
        layout.addWidget(self.scan_btn)

        self.confirm_all_btn = QPushButton("全部确认", bar)
        self.confirm_all_btn.clicked.connect(self._on_confirm_all)
        self.confirm_all_btn.setStyleSheet(
            "QPushButton { background: #107C10; color: white; border: 1px solid #107C10; }"
            "QPushButton:hover { background: #0E6A0E; }"
        )
        layout.addWidget(self.confirm_all_btn)

        self.batch_confirm_btn = QPushButton("批量确认已匹配", bar)
        self.batch_confirm_btn.clicked.connect(self._on_batch_confirm_matched)
        layout.addWidget(self.batch_confirm_btn)

        self.rematch_btn = QPushButton("重新匹配", bar)
        self.rematch_btn.clicked.connect(self._on_rematch)
        layout.addWidget(self.rematch_btn)

        self.export_btn = QPushButton("导出 Excel", bar)
        self.export_btn.clicked.connect(self._on_export_excel)
        layout.addWidget(self.export_btn)

        # v1.3.1：选择项目文件夹
        self.select_folder_btn = QPushButton("选择项目文件夹", bar)
        self.select_folder_btn.clicked.connect(self._on_select_project_folder)
        self.select_folder_btn.setStyleSheet(
            "QPushButton { background: #107C10; color: white; border: 1px solid #107C10; }"
            "QPushButton:hover { background: #0E6A0E; }"
        )
        layout.addWidget(self.select_folder_btn)

        # v1.6.3：整理文件按钮（预览 + 执行合并）
        self.organize_btn = QPushButton("整理文件", bar)
        self.organize_btn.clicked.connect(self._on_organize)
        self.organize_btn.setStyleSheet(
            "QPushButton { background: #0067C0; color: white; border: 1px solid #0067C0; }"
            "QPushButton:hover { background: #0B7AD9; }"
        )
        layout.addWidget(self.organize_btn)

        self.content_layout.addWidget(bar)

    # ------------------------------------------------------------------ 统计卡片

    def _setup_stat_cards(self) -> None:
        """统计卡片行（v1.2.2 增加已确认）。"""
        self.stat_cards = StatCardRow(self)
        self.stat_cards.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed
        )
        self.content_layout.addWidget(self.stat_cards)

    # ------------------------------------------------------------------ 结果区域

    def _setup_result_area(self) -> None:
        """结果区域：筛选栏 + 主内容（QSplitter）。"""
        self.filter_bar = ScanFilterBar(self)
        self.filter_bar.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed
        )
        self.filter_bar.filter_changed.connect(self._on_filter_changed)
        self.content_layout.addWidget(self.filter_bar)

        splitter = QSplitter(Qt.Orientation.Horizontal, self)
        splitter.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )

        self.result_table = ScanResultTable(splitter)
        self.result_table.item_selected.connect(self._on_item_selected)
        self.result_table.confirm_toggled.connect(self._on_confirm_toggled)
        splitter.addWidget(self.result_table)

        self.preview_panel = DetailPreviewPanel(splitter)
        splitter.addWidget(self.preview_panel)

        splitter.setStretchFactor(0, 7)
        splitter.setStretchFactor(1, 3)

        self.content_layout.addWidget(splitter, 1)

    # ------------------------------------------------------------------ 数据载入

    def load_project(self, project_id: int) -> bool:
        """载入项目并执行扫描（v1.2 旧入口，保留兼容）。

        v1.4 警告：此方法会触发扫描。新代码应使用 load_project_cached。
        """
        conn = Database.open_db_connection(self._config.database.path)
        try:
            projects_repository.init_projects_table(conn)
            projects_repository.init_point_dictionary_table(conn)
            row = projects_repository.fetch_project_by_id(conn, project_id)
        finally:
            conn.close()

        if row is None:
            logger.warning("扫描中心：项目 id=%s 不存在", project_id)
            self._project_id = None
            self._has_valid_session = False
            self._clear_display()
            return False

        self._project_id = project_id
        self._project_name = row["project_name"] or f"项目 #{project_id}"
        self._select_project_in_combo(project_id)
        self._load_and_scan()
        return True

    def load_project_cached(self, project_id: int) -> bool:
        """v1.4：仅加载缓存扫描结果（不触发扫描）。

        用于项目打开时快速显示已有结果，避免卡顿。
        """
        conn = Database.open_db_connection(self._config.database.path)
        try:
            projects_repository.init_projects_table(conn)
            projects_repository.init_point_dictionary_table(conn)
            row = projects_repository.fetch_project_by_id(conn, project_id)
        finally:
            conn.close()

        if row is None:
            logger.warning("扫描中心：项目 id=%s 不存在", project_id)
            self._project_id = None
            self._has_valid_session = False
            self._clear_display()
            return False

        self._project_id = project_id
        self._project_name = row["project_name"] or f"项目 #{project_id}"
        self._select_project_in_combo(project_id)

        # v1.5.3：优先读取当前 Scan Session；无有效 session 时不自动扫描
        from ....core.scan_controller import load_current_scan_session

        conn = Database.open_db_connection(self._config.database.path)
        try:
            session = load_current_scan_session(conn, project_id)
        finally:
            conn.close()

        if session is not None:
            items = _rebuild_items_from_dict(session.get("scan_result", {}))
            self._summary = ScanResultSummary.from_items(
                items=items,
                project_id=project_id,
                project_name=self._project_name,
                scan_directory=session.get("scan_path", ""),
                scan_duration_ms=session.get("scan_duration_ms", 0),
            )
            self._summary.scan_time = session.get("scan_time", self._summary.scan_time)
            self._has_valid_session = True
            self._display_summary()
            logger.info("扫描中心：从 Scan Session 加载项目 id=%s，%d 条结果", project_id, len(items))
        else:
            self._has_valid_session = False
            self._clear_display()
            self._refresh_project_combo()
            logger.info("扫描中心：项目 id=%s 无有效 Scan Session", project_id)

        return True

    def _load_and_scan(self) -> None:
        """v1.4 升级：使用 ScanController 统一扫描入口。

        加载点位字典 + 执行文件扫描 + 生成结果 + 写入数据库。
        """
        if self._project_id is None:
            self._has_valid_session = False
            self._clear_display()
            return

        conn = Database.open_db_connection(self._config.database.path)
        try:
            points = projects_repository.fetch_points_with_status(
                conn, self._project_id
            )
        finally:
            conn.close()

        if not points:
            logger.info("扫描中心：项目 id=%s 无点位数据", self._project_id)
            self._has_valid_session = False
            self._clear_display()
            self._refresh_project_combo()
            return

        # v1.4：使用 ScanController.run_scan 统一入口（含数据库写入）
        from ....core.scan_controller import ScanController
        from ....core.scan_result import ScanResultSummary

        try:
            result_dict = ScanController.run_scan(
                project_id=self._project_id,
                project_name=self._project_name,
                points=points,
                scan_directory="",
                db_path=self._config.database.path,
            )
            # 从 dict 重建 ScanResultSummary（用于 UI 渲染）
            items = _rebuild_items_from_dict(result_dict)
            self._summary = ScanResultSummary.from_items(
                items=items,
                project_id=self._project_id,
                project_name=self._project_name,
                scan_directory=result_dict.get("scan_directory", ""),
                scan_duration_ms=result_dict.get("scan_duration_ms", 0),
            )
        except Exception as exc:
            logger.exception("ScanController 扫描失败")
            QMessageBox.critical(self, "扫描失败", f"扫描过程中发生错误：\n{exc}")
            return

        self._has_valid_session = True
        self._display_summary()

    def _display_summary(self) -> None:
        """渲染扫描结果摘要到 UI。"""
        if self._summary is None:
            self._clear_display()
            return

        s = self._summary

        self._refresh_project_combo()
        self._select_project_in_combo(self._project_id)
        if s.scan_time:
            try:
                dt = datetime.fromisoformat(s.scan_time)
                self.scan_time_label.setText(
                    f"扫描时间：{dt.strftime('%Y-%m-%d %H:%M:%S')}"
                )
            except (ValueError, TypeError):
                self.scan_time_label.setText(f"扫描时间：{s.scan_time}")
        self.scan_duration_label.setText(f"耗时：{s.scan_duration_ms}ms")
        self.scan_dir_label.setText(f"目录：{s.scan_directory or '沙盒目录'}")

        self.stat_cards.update_stats(s)
        self.result_table.load_results(s.items)
        self._on_filter_changed()
        self.preview_panel.clear_detail()

        logger.info(
            "扫描中心渲染完成（v1.2.2）：总=%d 已匹配=%d 未找到=%d 已确认=%d",
            s.total_points, s.matched_count, s.not_found_count, s.confirmed_count,
        )
        self._update_scan_button_text()

    # ------------------------------------------------------------------ 项目选择器

    def _refresh_project_combo(self) -> None:
        """从数据库加载所有已导入项目，填充到下拉框中。"""
        self.project_combo.blockSignals(True)
        self.project_combo.clear()
        self.project_combo.addItem("— 请选择项目 —", None)

        try:
            conn = Database.open_db_connection(self._config.database.path)
            try:
                projects_repository.init_projects_table(conn)
                rows = projects_repository.fetch_all_projects(conn)
            finally:
                conn.close()
        except Exception as exc:
            logger.warning("加载项目列表失败：%s", exc)
            rows = []

        for row in rows:
            pid = row["id"]
            pname = row["project_name"] or f"项目 #{pid}"
            self.project_combo.addItem(pname, pid)

        self.project_combo.blockSignals(False)

    def _select_project_in_combo(self, project_id: int | None) -> None:
        """在下拉框中选中指定项目。"""
        if project_id is None:
            self.project_combo.setCurrentIndex(0)
            return
        for i in range(self.project_combo.count()):
            if self.project_combo.itemData(i) == project_id:
                self.project_combo.setCurrentIndex(i)
                return

    def _on_project_selected(self, index: int) -> None:
        """用户在下拉框中选择了一个项目。"""
        if index <= 0:
            return  # 占位项「请选择项目」

        project_id = self.project_combo.itemData(index)
        if project_id is None or project_id == self._project_id:
            return  # 同一个项目，不重复加载

        logger.info("扫描中心：用户选择项目 id=%s", project_id)
        self.load_project_cached(project_id)

    # ------------------------------------------------------------------ 清空显示

    def _clear_display(self) -> None:
        """清空全部显示。"""
        self._refresh_project_combo()
        self.scan_time_label.setText("扫描时间：—")
        self.scan_duration_label.setText("耗时：—")
        self.scan_dir_label.setText("目录：—")
        self.stat_cards.update_stats(ScanResultSummary())
        self.result_table.setRowCount(0)
        self.preview_panel.clear_detail()
        self._summary = None
        self._update_scan_button_text()

    def _update_scan_button_text(self) -> None:
        """根据当前 Scan Session 生命周期更新扫描按钮文案。"""
        if hasattr(self, "scan_btn"):
            self.scan_btn.setText("重新扫描" if self._has_valid_session else "执行扫描")

    # ------------------------------------------------------------------ 扫描操作

    def _on_scan(self) -> None:
        """执行扫描。"""
        if self._project_id is None:
            return
        logger.info("扫描中心：重新扫描项目 id=%s", self._project_id)
        self._load_and_scan()

    # ------------------------------------------------------------------ 确认操作（v1.2.2）

    def _on_confirm_toggled(self, _row: int, item: ScanResultItem) -> None:
        """用户点击确认按钮 → 切换确认状态并保存到历史。"""
        if self._project_id is None or self._summary is None:
            return

        item.confirmed = not item.confirmed

        if item.confirmed:
            item.match_method = "manual"
            if item.matched_folder:
                self._save_match_history(item)
        else:
            # 取消确认：重新计算匹配状态
            item.match_method = "fuzzy"
            if item.match_score >= 0.85:
                item.match_status = MatchStatus.MATCHED
            elif item.match_score >= 0.70:
                item.match_status = MatchStatus.PARTIAL_MATCH
            else:
                item.match_status = MatchStatus.NOT_FOUND

        # 刷新统计和列表
        self._summary = ScanResultSummary.from_items(
            items=self._summary.items,
            project_id=self._project_id,
            project_name=self._project_name,
            scan_directory=self._summary.scan_directory,
            scan_duration_ms=self._summary.scan_duration_ms,
        )
        self._display_summary()

    def _on_confirm_all(self) -> None:
        """全部确认：确认所有已匹配但有文件夹的结果。"""
        if self._project_id is None or self._summary is None:
            return

        count = 0
        for item in self._summary.items:
            if not item.confirmed and item.matched_folder:
                item.confirmed = True
                item.match_method = "manual"
                self._save_match_history(item)
                count += 1

        if count > 0:
            self._refresh_summary()
            QMessageBox.information(self, "全部确认", f"已确认 {count} 个点位。")
        else:
            QMessageBox.information(self, "提示", "没有可确认的未确认点位。")

    def _on_batch_confirm_matched(self) -> None:
        """批量确认已匹配项：仅确认 MATCHED 状态的未确认项。"""
        if self._project_id is None or self._summary is None:
            return

        count = 0
        for item in self._summary.items:
            if (
                not item.confirmed
                and item.match_status == MatchStatus.MATCHED
                and item.matched_folder
            ):
                item.confirmed = True
                item.match_method = "manual"
                self._save_match_history(item)
                count += 1

        if count > 0:
            self._refresh_summary()
            QMessageBox.information(self, "批量确认", f"已确认 {count} 个已匹配点位。")
        else:
            QMessageBox.information(self, "提示", "没有可批量确认的未确认已匹配点位。")

    def _save_match_history(self, item: ScanResultItem) -> None:
        """保存单条匹配历史到数据库。"""
        if self._project_id is None or not item.matched_folder:
            return
        try:
            conn = Database.open_db_connection(self._config.database.path)
            try:
                scan_match_history_repository.init_scan_match_history_table(conn)
                scan_match_history_repository.save_match_history(
                    conn,
                    self._project_id,
                    item.standard_point_name,
                    item.matched_folder,
                    item.match_method,
                )
            finally:
                conn.close()
        except Exception as exc:
            logger.warning("保存匹配历史失败：%s", exc)

    def _refresh_summary(self) -> None:
        """重新计算汇总统计并刷新 UI。"""
        if self._summary is None:
            return
        self._summary = ScanResultSummary.from_items(
            items=self._summary.items,
            project_id=self._project_id,
            project_name=self._project_name,
            scan_directory=self._summary.scan_directory,
            scan_duration_ms=self._summary.scan_duration_ms,
        )
        self._display_summary()

    # ------------------------------------------------------------------ 重新匹配（v1.2.2）

    def _on_rematch(self) -> None:
        """对选中行执行重新匹配。"""
        if self._project_id is None or self._summary is None:
            return

        row = self.result_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先在结果列表中点击选择一个点位。")
            return
        if row >= len(self._summary.items):
            return

        item = self._summary.items[row]
        # 只允许 NOT_FOUND / PARTIAL_MATCH / MULTIPLE_MATCH 重新匹配
        if item.match_status == MatchStatus.MATCHED and item.confirmed:
            QMessageBox.information(self, "提示", "已匹配且已确认的点位无需重新匹配。")
            return

        # 收集候选目录
        from ....core.scanner import TEST_ROOT_PATH, scan_project_root

        candidates: list[str] = []
        try:
            projects = scan_project_root(str(TEST_ROOT_PATH))
            from ....core.matcher import match_folder

            matched_project = None
            for proj in projects:
                result = match_folder(self._project_name, proj.name)
                if result.is_match:
                    matched_project = proj
                    break

            if matched_project is not None:
                for site in matched_project.sites:
                    candidates.append(site.name)
        except Exception as exc:
            logger.warning("获取候选目录失败：%s", exc)

        # 打开选择对话框
        dialog = RematchDialog(item.standard_point_name, candidates, self)
        if not dialog.exec():
            return  # 用户取消

        selected = dialog.selected_folder()
        if not selected:
            return

        # 更新匹配结果
        item.matched_folder = selected
        item.match_status = MatchStatus.MATCHED
        item.match_score = 1.0
        item.match_method = "manual"
        item.confirmed = True
        item.suggestion = f"已确认（manual）—— 用户重新匹配至「{selected}」"

        # 保存历史
        self._save_match_history(item)

        # 刷新
        self._refresh_summary()
        QMessageBox.information(
            self, "重新匹配完成",
            f"点位「{item.standard_point_name}」已重新匹配至：\n{selected}",
        )

    # ------------------------------------------------------------------ 导出 Excel（v1.2.2）

    def _on_export_excel(self) -> None:
        """导出扫描结果为 Excel 文件。"""
        if self._summary is None:
            QMessageBox.warning(self, "提示", "无扫描结果可导出。")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "导出扫描结果",
            f"扫描结果_{self._project_name}.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not path:
            return

        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "错误", "需要安装 openpyxl 库。")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "扫描结果"

            # 表头
            headers = [
                "序号", "标准点位名称", "实际目录", "区县",
                "CAD状态", "预算状态", "匹配率(%)", "匹配状态",
                "确认状态", "匹配方式", "建议",
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)

            # 数据行
            for i, item in enumerate(self._summary.items, 1):
                row_data = [
                    i,
                    item.standard_point_name,
                    item.matched_folder or "",
                    item.county,
                    item.cad_status,
                    item.budget_status,
                    item.match_percent,
                    item.match_status.label,
                    item.confirmed_label,
                    item.match_method,
                    item.suggestion,
                ]
                for col, val in enumerate(row_data, 1):
                    ws.cell(row=i + 1, column=col, value=val)

            wb.save(path)
            QMessageBox.information(
                self, "导出完成",
                f"已导出 {len(self._summary.items)} 条扫描结果到：\n{path}",
            )
            logger.info("扫描结果已导出到：%s（%d 条）", path, len(self._summary.items))

        except Exception as exc:
            logger.exception("导出 Excel 失败")
            QMessageBox.critical(self, "导出失败", f"导出 Excel 时出错：\n{exc}")

    # ------------------------------------------------------------------ 筛选

    def _on_filter_changed(self) -> None:
        f = self.filter_bar.filters()
        self.result_table.apply_filter(
            status_filter=f["status"],
            search_text=f["search"],
            cad_filter=f["cad"],
            budget_filter=f["budget"],
        )

    # ------------------------------------------------------------------ 详情预览

    def _on_item_selected(self, item: ScanResultItem) -> None:
        self.preview_panel.show_detail(item)

    # ------------------------------------------------------------------ v1.3.1 项目文件夹

    def _on_select_project_folder(self) -> None:
        """选择当前项目对应的文件夹。"""
        if self._project_id is None:
            QMessageBox.warning(self, "提示", "请先打开一个项目。")
            return

        folder = QFileDialog.getExistingDirectory(
            self, "选择项目文件夹", "",
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontResolveSymlinks,
        )
        if not folder:
            return

        try:
            from ....core import project_profile_repository as ppr
            from ....core.database import Database
            from ....core.scan_controller import invalidate_scan_session
            conn = Database.open_db_connection(self._config.database.path)
            try:
                ppr.init_project_profiles_table(conn)
                ppr.set_project_folder(conn, self._project_id, folder)
                invalidate_scan_session(conn, self._project_id)
            finally:
                conn.close()
        except Exception as exc:
            QMessageBox.critical(self, "错误", f"保存失败：{exc}")
            return

        self.scan_dir_label.setText(f"目录：{folder}")
        self._has_valid_session = False
        self._summary = None
        self.result_table.setRowCount(0)
        self.preview_panel.clear_detail()
        self._update_scan_button_text()
        QMessageBox.information(self, "已设置", f"项目文件夹：\n{folder}\n\n请点击「执行扫描」更新结果。")

    # ------------------------------------------------------------------ v1.3 文件整理

    def _build_organize_plan_from_current_session(self):
        """从当前 Scan Session 构建整理计划；不扫描、不重新归属。"""
        if self._project_id is None or not self._has_valid_session or self._summary is None:
            return None

        from ....core.scan_controller import load_current_scan_session
        from ....core.file_organizer import build_organize_plan_from_scan_session

        conn = Database.open_db_connection(self._config.database.path)
        try:
            session = load_current_scan_session(conn, self._project_id)
        finally:
            conn.close()

        if session is None:
            self._has_valid_session = False
            self._update_scan_button_text()
            return None

        points = [
            {"id": item.point_id, "standard_point_name": item.standard_point_name,
             "county": item.county}
            for item in self._summary.items
            if item.point_id is not None
        ]
        if not points:
            return None

        ownership = session.get("ownership", {})
        point_files = ownership.get("point_files", {})
        return build_organize_plan_from_scan_session(
            point_files=point_files,
            points=points,
            project_path=session.get("scan_path", ""),
        )

    def _on_organize(self) -> None:
        """整理文件：先弹出预览，确认后执行整理。"""
        if self._project_id is None or self._summary is None or not self._has_valid_session:
            QMessageBox.warning(self, "提示", "请先执行扫描")
            return

        try:
            plan = self._build_organize_plan_from_current_session()
            if plan is None:
                QMessageBox.warning(self, "提示", "请先执行扫描")
                return

            if plan.total_files == 0:
                QMessageBox.information(self, "提示", "未找到可整理的文件。")
                return

            # 弹出预览对话框（含「整理」+「取消」按钮）
            dialog = OrganizePreviewDialog(plan, self)
            if not dialog.exec():
                return  # 用户点取消

            # 用户点「整理」→ 执行整理
            from ....core.file_organizer import apply_organize_plan, cleanup_empty_dirs
            from ....core.scan_controller import invalidate_scan_session

            result = apply_organize_plan(plan)
            if result.get("moved", 0) > 0:
                conn = Database.open_db_connection(self._config.database.path)
                try:
                    invalidate_scan_session(conn, self._project_id)
                finally:
                    conn.close()
                self._has_valid_session = False
                self._update_scan_button_text()

            # 整理后自动清理项目目录下的空文件夹
            deleted = 0
            if plan.project_path:
                deleted = cleanup_empty_dirs(plan.project_path)

            QMessageBox.information(
                self, "整理完成",
                f"移动 {result['moved']} 个文件\n"
                f"跳过 {result['skipped']} 个文件\n"
                f"错误 {len(result['errors'])} 个\n"
                f"清理空文件夹 {deleted} 个\n\n"
                f"文件位置可能已变化，请点击「执行扫描」更新结果。",
            )

        except Exception as exc:
            logger.exception("文件整理失败")
            QMessageBox.critical(self, "错误", f"整理失败：{exc}")


# ====================================================================
# 辅助函数
# ====================================================================


def _rebuild_items_from_dict(result_dict: dict) -> list:
    """v1.4：从 ScanController.run_scan 返回的 dict 重建 ScanResultItem 列表。"""
    from ....core.scan_result import MatchStatus, ScanResultItem

    items = []
    for item_dict in result_dict.get("items", []):
        match_status = MatchStatus.NOT_FOUND
        try:
            match_status = MatchStatus[item_dict.get("match_status", "NOT_FOUND")]
        except KeyError:
            pass

        item = ScanResultItem(
            point_id=item_dict.get("point_id"),
            standard_point_name=item_dict.get("standard_point_name", ""),
            original_name=item_dict.get("original_name", ""),
            county=item_dict.get("county", ""),
            matched_folder=item_dict.get("matched_folder"),
            matched_folder_path=item_dict.get("matched_folder_path"),
            match_score=item_dict.get("match_score", 0.0),
            match_status=match_status,
            cad_status=item_dict.get("cad_status", "无"),
            budget_status=item_dict.get("budget_status", "无"),
            cad_file_count=item_dict.get("cad_file_count", 0),
            budget_file_count=item_dict.get("budget_file_count", 0),
            suggestion=item_dict.get("suggestion", ""),
            confirmed=item_dict.get("confirmed", False),
            match_method=item_dict.get("match_method", "fuzzy"),
            file_owner_point_id=item_dict.get("file_owner_point_id"),
            match_confidence=item_dict.get("match_confidence", 0.0),
            scanned_files=item_dict.get("scanned_files", []),
            dynamic_data=item_dict.get("dynamic_data", {}),
        )
        items.append(item)
    return items
