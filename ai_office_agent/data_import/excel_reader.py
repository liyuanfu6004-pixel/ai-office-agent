"""Excel 读取与表头自动识别模块。

使用 openpyxl 读取 .xlsx 文件。表头不固定，因此：
1. 在文件头部若干行内扫描，找到第一行"足够像个表头"的行作为表头行；
2. 该行之下所有连续非空行视为数据行。

判定规则详见 `_looks_like_header`，兼顾常见通信设计表格的列名。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ..utils.logger import setup_logger

logger = setup_logger()

# 表头扫描最大行数：只在前 N 行里找表头
_MAX_HEADER_SCAN_ROWS = 20

# 识别某个单元格是否属于"项目相关表头"的关键词（小写匹配）
_HEADER_KEYWORDS = (
    "项目名称",
    "项目类型",
    "年份",
    "年度",
    "区县",
    "区县数",
    "区县数量",
    "点位",
    "点位数",
    "点位数量",
    "状态",
)

# 单行至少命中多少个关键词才认为是表头行
_MIN_HEADER_HITS = 2

# 识别"子表头/合并标题"的关键词
# 子表头行不应被误判为表头起点，应继续往下找。
_CHILD_HEADER_KEYWORDS = (
    "立项批复", "设计批复", "结算", "批复",
    "时间", "规模", "金额", "文号",
    "工作量", "工时", "剩余",
)


def load_workbook(path: str | Path):
    """打开 Excel 工作簿（只读模式，省内存）。

    Args:
        path: .xlsx 文件路径。

    Returns:
        openpyxl Workbook 对象。
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {p}")

    wb = openpyxl.load_workbook(filename=str(p), read_only=True, data_only=True)
    return wb


def read_sheet(path: str | Path, sheet_name: str | None = None):
    """读取 Excel 文件并返回表头与数据行。

    动态识别表头行：在文件前若干行中扫描第一个命中足够关键词的行。
    若识别失败，则回退到"首行即表头"。

    Args:
        path: .xlsx 文件路径。
        sheet_name: 指定工作表名；None 则取第一个可见工作表。

    Returns:
        (headers, data_rows)
        - headers: 表头文本列表（已去 None，按原列顺序）。
          同时返回列号映射，见 `resolve_columns`。
        - data_rows: 行字典列表，键为表头文本，值为单元格原始值。

    Raises:
        ValueError: 文件中没有任何工作表或工作表为空。
    """
    wb = load_workbook(path)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            ws = wb[sheet_name]
        else:
            # 取第一个工作表
            ws = wb[wb.sheetnames[0]]

        # 提取所有行（迭代器），read_only 模式下只能迭代一次
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError("Excel 工作表为空")

    # 1. 动态识别表头行（v1.2.4：支持多层表头，返回最后一层表头行号 + 合并表头）
    header_idx, merged_headers = _build_merged_headers(rows)

    # 清洗表头：去 None，统一为字符串并去空白；空表头用"列N"兜底（v1.2.1：禁止 __col_x__）
    cleaned_headers = []
    for i, h in enumerate(merged_headers):
        if h is None or str(h).strip() == "":
            cleaned_headers.append(f"列{i + 1}")
        else:
            cleaned_headers.append(str(h).strip())

    # 2. 表头行以下为数据；每行转为 {表头: 单元格值} 字典
    data_rows: list[dict] = []
    for r in rows[header_idx + 1 :]:
        # 整行全空则跳过
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue
        row_dict = {}
        for col_idx, header in enumerate(cleaned_headers):
            row_dict[header] = r[col_idx] if col_idx < len(r) else None
        data_rows.append(row_dict)

    logger.info(
        "Excel 读取完成：%s，表头行=%d，列数=%d，数据行数=%d",
        Path(path).name,
        header_idx + 1 if header_idx is not None else 1,
        len(cleaned_headers),
        len(data_rows),
    )
    return cleaned_headers, data_rows


def _build_merged_headers(rows: list[tuple]) -> tuple[int, list]:
    """识别主表头 + 合并子表头，返回 (最后一层表头行号, 合并后的表头列表)。

    合并规则（按列向下找第一个非空值，作为该列表头）：
        - 主表头第 0 列="序号"，子表头第 0 列=None → 合并表头第 0 列="序号"
        - 主表头第 11 列=None，子表头第 11 列="立项批复\n时间" → 合并表头第 11 列="立项批复\n时间"

    v1.2.4 升级：支持多层表头。
    """
    from ..core.matcher import match_field

    scan_limit = min(len(rows), _MAX_HEADER_SCAN_ROWS)

    # 第一步：找主表头行
    main_idx = None
    for idx in range(scan_limit):
        row = rows[idx] or ()
        hits = 0
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            if any(match_field(text, kw).is_match for kw in _HEADER_KEYWORDS):
                hits += 1
        if hits >= _MIN_HEADER_HITS:
            main_idx = idx
            break

    if main_idx is None:
        # 回退：第 0 行为表头
        return 0, list(rows[0]) if rows else []

    # 第二步：扫描主表头后是否紧跟子表头行
    last_header_idx = main_idx
    for idx in range(main_idx + 1, min(main_idx + 6, len(rows))):
        row = rows[idx] or ()
        if _looks_like_data_row(row):
            break
        if _looks_like_child_header_row(row):
            last_header_idx = idx
        else:
            break

    # 第三步：合并主表头 + 各层子表头（按列向下取第一个非空）
    header_rows = [rows[i] or () for i in range(main_idx, last_header_idx + 1)]
    if not header_rows:
        return main_idx, list(rows[main_idx]) if main_idx < len(rows) else []

    n_cols = max(len(r) for r in header_rows)
    merged: list = [None] * n_cols
    for col in range(n_cols):
        for r in header_rows:
            if col < len(r) and r[col] is not None and str(r[col]).strip():
                merged[col] = r[col]
                break

    return last_header_idx, merged


def _detect_header_row(rows: list[tuple]) -> int | None:
    """向后兼容：仅返回主表头行号（最后一层表头位置）。"""
    from ..core.matcher import match_field

    scan_limit = min(len(rows), _MAX_HEADER_SCAN_ROWS)
    main_idx = None
    for idx in range(scan_limit):
        row = rows[idx] or ()
        hits = 0
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            if any(match_field(text, kw).is_match for kw in _HEADER_KEYWORDS):
                hits += 1
        if hits >= _MIN_HEADER_HITS:
            main_idx = idx
            break
    if main_idx is None:
        return None
    last_header_idx = main_idx
    for idx in range(main_idx + 1, min(main_idx + 6, len(rows))):
        row = rows[idx] or ()
        if _looks_like_data_row(row):
            break
        if _looks_like_child_header_row(row):
            last_header_idx = idx
        else:
            break
    return last_header_idx


def _looks_like_data_row(row: tuple) -> bool:
    """判断一行是否更像"数据行"而非子表头。

    数据特征：含连续数字/编码、含长字符串项目名、含 MIS 项目编号格式（如 C 开头+数字）。
    """
    if not row:
        return False
    text_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
    if not text_cells:
        return False
    # 含 MIS 项目编号（C 字母+长数字）→ 必是数据行
    import re as _re
    for t in text_cells:
        if _re.search(r"\bC\d{10,}\b", t) or _re.search(r"\b\d{6,}\b", t):
            return True
        # 含"中国移动"项目名特征
        if "中国移动" in t and len(t) > 20:
            return True
    return False


def _looks_like_child_header_row(row: tuple) -> bool:
    """判断一行是否像"子表头"（多层表头中的下一级）。

    子表头特征：含合并标题的子列关键词（时间/规模/金额等），但不含业务数据。
    """
    from ..core.matcher import match_field
    if not row:
        return False
    # 行内至少 2 个非空单元格命中子表头关键词
    hits = 0
    for cell in row:
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        if any(match_field(text, kw).is_match for kw in _CHILD_HEADER_KEYWORDS):
            hits += 1
    return hits >= 2
